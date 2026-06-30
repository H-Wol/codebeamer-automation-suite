from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
import pandas as pd

from src.gui.services import GuiCodebeamerService
from src.gui.services import GuiExcelService
from src.gui.services import GuiUploadPipelineService
from src.gui.settings_store import GuiSettings
from src.models import MappingStatus
from src.models import PayloadStatus


class FakeExcelReader:
    def __init__(self, header_row: int = 1, summary_col: str = "Summary", logger=None) -> None:
        del logger
        self.header_row = header_row
        self.summary_col = summary_col

    def list_sheet_names(self, file_path: str) -> list[str]:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            return wb.sheetnames
        finally:
            wb.close()

    def read_headers(self, file_path: str, sheet_name: str | int) -> list[str]:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            row = next(ws.iter_rows(min_row=self.header_row, max_row=self.header_row, values_only=True), ())
            return [str(value).strip() if value is not None else f"Unnamed_{index}" for index, value in enumerate(row)]
        finally:
            wb.close()

    def read_excel(self, file_path: str, sheet_name: str | int = 0, visible: bool = False) -> pd.DataFrame:
        del visible
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            headers = self.read_headers(file_path, sheet_name)
            summary_index = headers.index(self.summary_col)
            records = []
            for excel_row, values in enumerate(rows[self.header_row:], start=self.header_row + 1):
                normalized = list(values)
                if all(value is None or str(value).strip() == "" for value in normalized[:len(headers)]):
                    continue
                record = {header: normalized[index] if index < len(normalized) else None for index, header in enumerate(headers)}
                record["_excel_row"] = excel_row
                record["_summary_indent"] = 0 if normalized[summary_index] is not None else 0
                records.append(record)
            return pd.DataFrame(records)
        finally:
            wb.close()


class FakeClient:
    def __init__(self, base_url, username, password, logger=None, **kwargs) -> None:
        del logger, kwargs
        self.base_url = base_url
        self.username = username
        self.password = password

    def get_projects(self):
        return [
            {"id": 10, "name": "Project A"},
            {"id": 20, "name": "Project B"},
        ]

    def get_trackers(self, project_id: int):
        return [
            {"id": project_id * 100, "name": f"Tracker {project_id}-1"},
            {"id": project_id * 100 + 1, "name": f"Tracker {project_id}-2"},
        ]

    def get_tracker_schema(self, tracker_id: int):
        del tracker_id
        return {
            "fields": [
                {
                    "id": 1,
                    "name": "Summary",
                    "type": "TextField",
                    "trackerItemField": "name",
                    "valueModel": "TextFieldValue",
                    "multipleValues": False,
                },
                {
                    "id": 2,
                    "name": "Status",
                    "type": "OptionChoiceField",
                    "trackerItemField": "status",
                    "valueModel": "ChoiceFieldValue<ChoiceOptionReference>",
                    "multipleValues": False,
                    "options": [
                        {"id": 201, "name": "Open", "type": "ChoiceOptionReference"},
                        {"id": 202, "name": "Review", "type": "ChoiceOptionReference"},
                    ],
                },
                {
                    "id": 3,
                    "name": "담당자",
                    "type": "TextField",
                    "trackerItemField": None,
                    "valueModel": "TextFieldValue",
                    "multipleValues": False,
                },
                {
                    "id": 4,
                    "name": "id",
                    "type": "TextField",
                    "trackerItemField": "id",
                    "valueModel": "TextFieldValue",
                    "multipleValues": False,
                },
                {
                    "id": 5,
                    "name": "parent",
                    "type": "TextField",
                    "trackerItemField": "parent",
                    "valueModel": "TextFieldValue",
                    "multipleValues": False,
                },
                {
                    "id": 6,
                    "name": "테이블필드",
                    "type": "TableField",
                    "trackerItemField": None,
                    "valueModel": "TableFieldValue",
                    "multipleValues": False,
                    "columns": [
                        {
                            "id": 31,
                            "name": "컬럼A",
                            "type": "TextField",
                            "valueModel": "TextFieldValue",
                        }
                    ],
                },
            ]
        }

    def create_item(self, tracker_id: int, payload: dict, parent_item_id: int | None = None):
        del tracker_id, payload, parent_item_id
        return {"id": 1}


class TrackerItemQueryFakeClient(FakeClient):
    all_search_calls: list[tuple[int, str]] = []

    def __init__(self, base_url, username, password, logger=None, **kwargs) -> None:
        super().__init__(base_url, username, password, logger=logger, **kwargs)
        self.search_calls: list[tuple[int, str]] = []

    def get_tracker_schema(self, tracker_id: int):
        schema = super().get_tracker_schema(tracker_id)
        schema["fields"].append({
            "id": 7,
            "name": "연관 요구사항",
            "type": "TrackerItemChoiceField",
            "valueModel": "ChoiceFieldValue<TrackerItemReference>",
            "multipleValues": True,
        })
        return schema

    def get_tracker_configuration(self, tracker_id: int):
        del tracker_id
        return {
            "basicInformation": {
                "trackerId": 1000,
                "name": "Test Tracker",
            },
            "fields": [
                {
                    "label": "연관 요구사항",
                    "choiceConfigOptionsSetting": {
                        "referenceFilters": [
                            {
                                "domainType": "TRACKER",
                                "domainId": 13526611,
                            }
                        ]
                    },
                }
            ],
        }

    def search_tracker_items_by_name(self, *, tracker_id: int, name: str, **kwargs):
        del kwargs
        self.search_calls.append((tracker_id, name))
        self.__class__.all_search_calls.append((tracker_id, name))
        lookup = {
            "REQ-100": [{"id": 9001, "name": "REQ-100", "type": "TrackerItemReference"}],
            "REQ-200": [{"id": 9002, "name": "REQ-200", "type": "TrackerItemReference"}],
        }
        return lookup.get(name, [])


class TrackerItemNonTrackerConfigFakeClient(TrackerItemQueryFakeClient):
    def get_tracker_configuration(self, tracker_id: int):
        del tracker_id
        return {
            "basicInformation": {
                "trackerId": 1000,
                "name": "Test Tracker",
            },
            "fields": [
                {
                    "label": "연관 요구사항",
                    "choiceConfigOptionsSetting": {
                        "referenceFilters": [
                            {
                                "domainType": "PROJECT",
                                "domainId": 13526611,
                            }
                        ]
                    },
                }
            ],
        }


class TrackerItemReferenceIdConfigFakeClient(TrackerItemQueryFakeClient):
    def get_tracker_schema(self, tracker_id: int):
        schema = super().get_tracker_schema(tracker_id)
        for field in schema["fields"]:
            if field.get("id") == 7:
                field["id"] = 17
                field["name"] = "SUDS 링크"
        return schema

    def get_tracker_configuration(self, tracker_id: int):
        del tracker_id
        return {
            "basicInformation": {
                "trackerId": 1000,
                "name": "Test Tracker",
            },
            "fields": [
                {
                    "referenceId": 17,
                    "label": "Software Unit Design Specification",
                    "choiceConfigOptionsSetting": {
                        "referenceFilters": [
                            {
                                "domainType": "TRACKER",
                                "domainId": 13526611,
                            }
                        ]
                    },
                }
            ],
        }


class UserReferenceDefaultFakeClient(FakeClient):
    def get_tracker_schema(self, tracker_id: int):
        schema = super().get_tracker_schema(tracker_id)
        schema["fields"].append({
            "id": 8,
            "name": "담당 사용자",
            "type": "ReferenceField",
            "referenceType": "UserReference",
            "valueModel": "ChoiceFieldValue<UserReference>",
            "multipleValues": False,
        })
        return schema


class CountingGuiExcelService(GuiExcelService):
    def __init__(self, **kwargs) -> None:
        super().__init__(**kwargs)
        self.load_preview_calls = 0

    def load_preview(self, *args, **kwargs):
        self.load_preview_calls += 1
        return super().load_preview(*args, **kwargs)


class CountingBatchExcelReader(FakeExcelReader):
    read_excel_calls: list[str] = []
    read_headers_calls: list[str] = []
    list_sheet_calls: list[str] = []

    @classmethod
    def reset_counts(cls) -> None:
        cls.read_excel_calls = []
        cls.read_headers_calls = []
        cls.list_sheet_calls = []

    def list_sheet_names(self, file_path: str) -> list[str]:
        self.__class__.list_sheet_calls.append(str(file_path))
        return super().list_sheet_names(file_path)

    def read_headers(self, file_path: str, sheet_name: str | int) -> list[str]:
        self.__class__.read_headers_calls.append(str(file_path))
        return super().read_headers(file_path, sheet_name)

    def read_excel(self, file_path: str, sheet_name: str | int = 0, visible: bool = False) -> pd.DataFrame:
        self.__class__.read_excel_calls.append(str(file_path))
        return super().read_excel(file_path, sheet_name=sheet_name, visible=visible)


class GuiCodebeamerServiceTest(unittest.TestCase):
    def test_connection_and_tracker_loading(self) -> None:
        service = GuiCodebeamerService(client_factory=FakeClient)
        settings = GuiSettings(
            base_url="https://example.com/cb",
            username="user",
            password="secret",
            rate_limit_retry_delay_seconds=1.0,
            rate_limit_max_retries=5,
        )

        projects = service.test_connection_and_load_projects(settings)
        trackers = service.load_trackers(settings, 10)

        self.assertEqual(projects[0]["id"], 10)
        self.assertEqual(projects[0]["name"], "Project A")
        self.assertEqual(trackers[0]["id"], 1000)
        self.assertEqual(trackers[1]["name"], "Tracker 10-2")


class GuiExcelServiceTest(unittest.TestCase):
    def test_load_preview_reads_sheet_names_headers_rows_and_summary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary", "담당자", "비고"])
            sheet.append(["REQ-001", "홍길동", "메모"])
            sheet.append(["REQ-002", "김철수", "메모2"])
            workbook.create_sheet("Other")
            workbook.save(path)
            workbook.close()

            preview = GuiExcelService(reader_cls=FakeExcelReader).load_preview(
                str(path),
                sheet_name="Main",
                header_row=1,
                max_preview_rows=5,
            )

            self.assertEqual(preview.sheet_names, ["Main", "Other"])
            self.assertEqual(preview.headers, ["Summary", "담당자", "비고"])
            self.assertEqual(preview.rows[0], ["REQ-001", "홍길동", "메모"])
            self.assertEqual(preview.suggested_summary, "Summary")

    def test_load_preview_preloads_raw_data_for_all_selected_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            first_path = Path(tmp_dir) / "first.xlsx"
            second_path = Path(tmp_dir) / "second.xlsx"
            for path, summary in (
                (first_path, "REQ-001"),
                (second_path, "REQ-002"),
            ):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Main"
                sheet.append(["Summary", "담당자"])
                sheet.append([summary, "홍길동"])
                workbook.save(path)
                workbook.close()

            CountingBatchExcelReader.reset_counts()
            preview = GuiExcelService(reader_cls=CountingBatchExcelReader).load_preview(
                str(first_path),
                file_paths=[str(first_path), str(second_path)],
                sheet_name="Main",
                header_row=1,
                summary_column="Summary",
            )

            self.assertEqual(
                sorted(preview.raw_df_by_file.keys()),
                sorted([str(first_path), str(second_path)]),
            )
            self.assertEqual(
                CountingBatchExcelReader.read_excel_calls,
                [str(first_path), str(second_path)],
            )


class GuiUploadPipelineServiceTest(unittest.TestCase):
    def test_prepare_mapping_context_builds_upload_columns_and_default_mapping(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary", "담당자", "테이블필드.컬럼A", "id", "parent"])
            sheet.append(["REQ-001", "홍길동", "값1", "1", ""])
            workbook.save(path)
            workbook.close()

            excel_service = GuiExcelService(reader_cls=FakeExcelReader)
            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=excel_service,
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )
            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )

            self.assertIn("Summary", mapping_context.upload_columns)
            self.assertIn("담당자", mapping_context.upload_columns)
            self.assertIn("테이블필드.컬럼A", mapping_context.upload_columns)
            self.assertNotIn("id", mapping_context.upload_columns)
            self.assertNotIn("parent", mapping_context.upload_columns)
            self.assertNotIn("id", mapping_context.schema_df["field_name"].tolist())
            self.assertNotIn("parent", mapping_context.schema_df["field_name"].tolist())
            self.assertEqual(mapping_context.selected_mapping["Summary"], "Summary")
            self.assertEqual(mapping_context.selected_mapping["담당자"], "담당자")
            self.assertEqual(mapping_context.selected_mapping["테이블필드.컬럼A"], "테이블필드")
            self.assertEqual(
                [candidate.schema_field for candidate in mapping_context.default_value_candidates],
                ["Summary", "Status", "담당자"],
            )
            self.assertTrue(mapping_context.default_value_candidates[0].allows_custom_value)
            self.assertEqual(mapping_context.default_value_candidates[1].options, ["Open", "Review"])
            self.assertEqual(mapping_context.file_paths, [str(path)])
            self.assertEqual(mapping_context.representative_file_path, str(path))

    def test_prepare_mapping_context_uses_file_selection_header_and_summary(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["메타", "메타", "메타"])
            sheet.append(["요약", "담당자", "비고"])
            sheet.append(["REQ-001", "홍길동", "메모"])
            workbook.save(path)
            workbook.close()

            excel_service = GuiExcelService(reader_cls=FakeExcelReader)
            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=excel_service,
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )

            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 2,
                    "summary_column": "요약",
                },
            )

            self.assertEqual(mapping_context.wizard.reader.header_row, 2)
            self.assertEqual(mapping_context.wizard.reader.summary_col, "요약")
            self.assertEqual(mapping_context.wizard.processor.summary_col, "요약")
            self.assertIn("요약", mapping_context.upload_columns)
            self.assertEqual(
                mapping_context.wizard.state.upload_df.iloc[0]["upload_name"],
                "REQ-001",
            )

    def test_prepare_mapping_context_keeps_original_upload_rows(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary", "담당자", "비고"])
            sheet.append(["REQ-001", "홍길동", "메모"])
            workbook.save(path)
            workbook.close()

            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=GuiExcelService(reader_cls=FakeExcelReader),
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )

            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )

            upload_df = mapping_context.wizard.state.upload_df

            self.assertEqual(list(upload_df["upload_name"]), ["REQ-001"])
            self.assertNotIn("_synthetic_root", upload_df.columns)

    def test_prepare_mapping_context_reuses_cached_preview_data(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary", "담당자"])
            sheet.append(["REQ-001", "홍길동"])
            workbook.save(path)
            workbook.close()

            excel_service = CountingGuiExcelService(reader_cls=FakeExcelReader)
            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=excel_service,
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )
            preview = excel_service.load_preview(
                str(path),
                sheet_name="Main",
                header_row=1,
                summary_column="Summary",
            )

            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "preview_file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                    "preview_data": preview,
                },
            )

            self.assertEqual(excel_service.load_preview_calls, 1)
            self.assertIsNotNone(mapping_context.preview_data)
            self.assertEqual(mapping_context.preview_data.file_path, str(path))

    def test_prepare_mapping_context_uses_tracker_configuration_for_query_defaults(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            TrackerItemQueryFakeClient.all_search_calls = []
            path = Path(tmp_dir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary", "연관 요구사항"])
            sheet.append(["REQ-001", "REQ-100"])
            workbook.save(path)
            workbook.close()

            service = GuiUploadPipelineService(
                client_factory=TrackerItemQueryFakeClient,
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )

            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "preview_file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )

            self.assertEqual(
                mapping_context.selected_tracker_item_settings["연관 요구사항"]["mode"],
                "query",
            )
            self.assertEqual(
                mapping_context.selected_tracker_item_settings["연관 요구사항"]["source_tracker_ids"],
                [13526611],
            )

    def test_prepare_mapping_context_uses_tracker_configuration_reference_id_when_name_differs(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary", "SUDS 링크"])
            sheet.append(["REQ-001", "REQ-100"])
            workbook.save(path)
            workbook.close()

            service = GuiUploadPipelineService(
                client_factory=TrackerItemReferenceIdConfigFakeClient,
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )

            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "preview_file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )

            self.assertEqual(
                mapping_context.selected_tracker_item_settings["SUDS 링크"]["mode"],
                "query",
            )
            self.assertEqual(
                mapping_context.selected_tracker_item_settings["SUDS 링크"]["source_tracker_ids"],
                [13526611],
            )
            tracker_field_row = mapping_context.schema_df[
                mapping_context.schema_df["field_name"] == "SUDS 링크"
            ].iloc[0]
            self.assertEqual(int(tracker_field_row["field_id"]), 17)
            self.assertEqual(tracker_field_row["tracker_item_source_tracker_ids"], [13526611])

    def test_prepare_mapping_context_disables_query_for_non_tracker_configuration(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary", "연관 요구사항"])
            sheet.append(["REQ-001", "REQ-100"])
            workbook.save(path)
            workbook.close()

            service = GuiUploadPipelineService(
                client_factory=TrackerItemNonTrackerConfigFakeClient,
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )

            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "preview_file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )

            self.assertEqual(
                mapping_context.selected_tracker_item_settings["연관 요구사항"]["mode"],
                "regex",
            )
            self.assertEqual(
                mapping_context.selected_tracker_item_settings["연관 요구사항"]["source_tracker_ids"],
                [],
            )
            tracker_item_candidate = next(
                candidate
                for candidate in mapping_context.tracker_item_field_candidates
                if candidate.schema_field == "연관 요구사항"
            )
            self.assertEqual(tracker_item_candidate.query_status, "unsupported")

    def test_prepare_mapping_context_includes_user_reference_field_in_default_value_candidates(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary"])
            sheet.append(["REQ-001"])
            workbook.save(path)
            workbook.close()

            service = GuiUploadPipelineService(
                client_factory=UserReferenceDefaultFakeClient,
                excel_service=GuiExcelService(reader_cls=FakeExcelReader),
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )

            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )

            default_candidates = {
                candidate.schema_field: candidate
                for candidate in mapping_context.default_value_candidates
            }
            self.assertIn("담당 사용자", default_candidates)
            self.assertTrue(default_candidates["담당 사용자"].allows_custom_value)

    def test_prime_tracker_item_lookup_cache_deduplicates_values_across_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            TrackerItemQueryFakeClient.all_search_calls = []
            path_a = Path(tmp_dir) / "a.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary", "연관 요구사항"])
            sheet.append(["REQ-001", "REQ-100"])
            workbook.save(path_a)
            workbook.close()

            path_b = Path(tmp_dir) / "b.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary", "연관 요구사항"])
            sheet.append(["REQ-002", "REQ-100"])
            sheet.append(["REQ-003", "REQ-200"])
            workbook.save(path_b)
            workbook.close()

            service = GuiUploadPipelineService(
                client_factory=TrackerItemQueryFakeClient,
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )

            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path_a),
                    "file_paths": [str(path_a), str(path_b)],
                    "preview_file_path": str(path_a),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )
            service.validate_mapping(
                mapping_context,
                {"Summary": "Summary", "연관 요구사항": "연관 요구사항"},
                selected_tracker_item_settings=mapping_context.selected_tracker_item_settings,
            )

            service._prime_tracker_item_lookup_cache_for_batch(settings, mapping_context)

            self.assertEqual(
                TrackerItemQueryFakeClient.all_search_calls,
                [(13526611, "REQ-100"), (13526611, "REQ-200")],
            )
            self.assertEqual(
                sorted(mapping_context.tracker_item_lookup_cache.keys()),
                [("연관 요구사항", "req-100"), ("연관 요구사항", "req-200")],
            )

    def test_build_root_item_preview_context_parses_named_groups(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "ABC_REQ-001.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary"])
            sheet.append(["REQ-001"])
            workbook.save(path)
            workbook.close()

            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=GuiExcelService(reader_cls=FakeExcelReader),
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )
            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "file_paths": [str(path)],
                    "preview_file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )

            preview_context = service.build_root_item_preview_context(
                mapping_context,
                {
                    "regex_pattern": r"^(?P<project>[A-Z]+)_(?P<title>.+)$",
                    "regex_target": "file_stem",
                    "field_sources": {"Summary": "title"},
                },
            )

            self.assertFalse(preview_context.has_blocking_issues)
            self.assertIn("project", preview_context.preview_columns)
            self.assertIn("title", preview_context.preview_columns)
            self.assertEqual(preview_context.preview_rows[0]["project"], "ABC")
            self.assertEqual(preview_context.preview_rows[0]["title"], "REQ-001")
            self.assertEqual(
                preview_context.field_assignments["Summary"],
                {
                    "enabled": True,
                    "mode": "file_source",
                    "value": "title",
                },
            )
            status_candidate = next(
                candidate
                for candidate in preview_context.field_candidates
                if candidate.schema_field == "Status"
            )
            summary_candidate = next(
                candidate
                for candidate in preview_context.field_candidates
                if candidate.schema_field == "Summary"
            )
            self.assertTrue(status_candidate.allows_fixed_value)
            self.assertEqual(status_candidate.fixed_options, ["Open", "Review"])
            self.assertTrue(summary_candidate.allows_fixed_value)
            self.assertTrue(summary_candidate.allows_custom_value)

    def test_build_root_item_payload_spec_uses_regex_mapped_name(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "ABC_REQ-001.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary"])
            sheet.append(["REQ-001"])
            workbook.save(path)
            workbook.close()

            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=GuiExcelService(reader_cls=FakeExcelReader),
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )
            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "file_paths": [str(path)],
                    "preview_file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )
            mapping_context.root_item_config = {
                "regex_pattern": r"^(?P<project>[A-Z]+)_(?P<title>.+)$",
                "regex_target": "file_stem",
                "field_assignments": {
                    "Summary": {
                        "enabled": True,
                        "mode": "file_source",
                        "value": "title",
                    },
                    "Status": {
                        "enabled": True,
                        "mode": "fixed_value",
                        "value": "Open",
                    },
                },
            }

            root_item_name, root_field_values = service.build_root_item_payload_spec(mapping_context, str(path))

            self.assertEqual(root_item_name, "REQ-001")
            self.assertEqual(root_field_values["Summary"], "REQ-001")
            self.assertEqual(root_field_values["Status"], "Open")

    def test_build_root_item_payload_spec_accepts_custom_scalar_fixed_value(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "ABC_REQ-001.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary"])
            sheet.append(["REQ-001"])
            workbook.save(path)
            workbook.close()

            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=GuiExcelService(reader_cls=FakeExcelReader),
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )
            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "file_paths": [str(path)],
                    "preview_file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )
            mapping_context.root_item_config = {
                "regex_pattern": "",
                "regex_target": "file_stem",
                "field_assignments": {
                    "담당자": {
                        "enabled": True,
                        "mode": "fixed_value",
                        "value": "홍길동",
                    },
                },
            }

            root_item_name, root_field_values = service.build_root_item_payload_spec(mapping_context, str(path))

            self.assertEqual(root_item_name, "ABC_REQ-001")
            self.assertEqual(root_field_values["담당자"], "홍길동")

    def test_build_root_item_preview_context_does_not_block_when_root_item_is_disabled(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "ABC_REQ-001.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary"])
            sheet.append(["REQ-001"])
            workbook.save(path)
            workbook.close()

            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=GuiExcelService(reader_cls=FakeExcelReader),
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )
            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "file_paths": [str(path)],
                    "preview_file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )

            preview_context = service.build_root_item_preview_context(
                mapping_context,
                {
                    "enabled": False,
                    "regex_pattern": "(",
                    "field_assignments": {
                        "Summary": {
                            "enabled": True,
                            "mode": "file_source",
                            "value": "missing",
                        }
                    },
                },
            )

            self.assertFalse(preview_context.enabled)
            self.assertFalse(preview_context.has_blocking_issues)
            self.assertEqual(preview_context.preview_columns, ["file_name"])
            self.assertIn("무시", preview_context.status_message)

    def test_build_root_item_payload_spec_returns_none_when_root_item_is_disabled(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "ABC_REQ-001.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary"])
            sheet.append(["REQ-001"])
            workbook.save(path)
            workbook.close()

            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=GuiExcelService(reader_cls=FakeExcelReader),
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )
            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "file_paths": [str(path)],
                    "preview_file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )
            mapping_context.root_item_config = {
                "enabled": False,
                "regex_pattern": r"^(?P<title>.+)$",
                "field_assignments": {
                    "Summary": {
                        "enabled": True,
                        "mode": "file_source",
                        "value": "title",
                    },
                },
            }

            root_item_name, root_field_values = service.build_root_item_payload_spec(mapping_context, str(path))

            self.assertIsNone(root_item_name)
            self.assertEqual(root_field_values, {})

    def test_prepare_mapping_context_rejects_mismatched_batch_headers(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            first_path = Path(tmp_dir) / "ABC_REQ-001.xlsx"
            second_path = Path(tmp_dir) / "DEF_REQ-002.xlsx"

            first_book = Workbook()
            first_sheet = first_book.active
            first_sheet.title = "Main"
            first_sheet.append(["Summary", "담당자", "비고"])
            first_sheet.append(["REQ-001", "홍길동", "메모"])
            first_book.save(first_path)
            first_book.close()

            second_book = Workbook()
            second_sheet = second_book.active
            second_sheet.title = "Main"
            second_sheet.append(["Summary", "상태", "비고"])
            second_sheet.append(["REQ-002", "Open", "메모2"])
            second_book.save(second_path)
            second_book.close()

            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=GuiExcelService(reader_cls=FakeExcelReader),
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )

            with self.assertRaisesRegex(ValueError, "헤더가 기준 파일과 다릅니다"):
                service.prepare_mapping_context(
                    settings,
                    {
                        "file_path": str(first_path),
                        "file_paths": [str(first_path), str(second_path)],
                        "preview_file_path": str(first_path),
                        "sheet_name": "Main",
                        "header_row": 1,
                        "summary_column": "Summary",
                    },
                )

    def test_run_batch_upload_aggregates_multiple_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            first_path = Path(tmp_dir) / "ABC_REQ-001.xlsx"
            second_path = Path(tmp_dir) / "DEF_REQ-002.xlsx"

            for path, summary in (
                (first_path, "REQ-001"),
                (second_path, "REQ-002"),
            ):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Main"
                sheet.append(["Summary"])
                sheet.append([summary])
                workbook.save(path)
                workbook.close()

            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=GuiExcelService(reader_cls=FakeExcelReader),
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )
            file_state = {
                "file_path": str(first_path),
                "file_paths": [str(first_path), str(second_path)],
                "preview_file_path": str(first_path),
                "sheet_name": "Main",
                "header_row": 1,
                "summary_column": "Summary",
            }

            mapping_context = service.prepare_mapping_context(settings, file_state)
            mapping_context.root_item_config = {
                "regex_pattern": r"^(?P<project>[A-Z]+)_(?P<title>.+)$",
                "regex_target": "file_stem",
                "field_assignments": {
                    "Summary": {
                        "enabled": True,
                        "mode": "file_source",
                        "value": "title",
                    },
                    "Status": {
                        "enabled": True,
                        "mode": "fixed_value",
                        "value": "Open",
                    },
                },
            }
            validation_context = service.validate_mapping(
                mapping_context,
                mapping_context.selected_mapping,
            )

            self.assertFalse(validation_context.has_blocking_issues)
            self.assertEqual(validation_context.summary_stats["file_count"], 2)
            self.assertEqual(validation_context.summary_stats["batch_total_rows"], 2)

            result = service.run_batch_upload(
                settings,
                file_state,
                mapping_context,
                dry_run=True,
                continue_on_error=True,
                output_dir=str(Path(tmp_dir) / "output"),
            )

            success_df = result["success_df"]
            self.assertEqual(len(success_df), 4)
            self.assertEqual(set(success_df["source_file"].tolist()), {"ABC_REQ-001.xlsx", "DEF_REQ-002.xlsx"})
            root_rows = success_df[success_df["_row_id"].isna()].reset_index(drop=True)
            self.assertEqual(root_rows["upload_name"].tolist(), ["REQ-001", "REQ-002"])
            self.assertTrue(result["failed_df"].empty)
            self.assertTrue(result["unresolved_df"].empty)

    def test_run_batch_upload_reuses_preloaded_raw_data_for_all_files(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            first_path = Path(tmp_dir) / "ABC_REQ-001.xlsx"
            second_path = Path(tmp_dir) / "DEF_REQ-002.xlsx"

            for path, summary in (
                (first_path, "REQ-001"),
                (second_path, "REQ-002"),
            ):
                workbook = Workbook()
                sheet = workbook.active
                sheet.title = "Main"
                sheet.append(["Summary"])
                sheet.append([summary])
                workbook.save(path)
                workbook.close()

            CountingBatchExcelReader.reset_counts()
            excel_service = GuiExcelService(reader_cls=CountingBatchExcelReader)
            preview = excel_service.load_preview(
                str(first_path),
                file_paths=[str(first_path), str(second_path)],
                sheet_name="Main",
                header_row=1,
                summary_column="Summary",
            )

            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=excel_service,
                reader_cls=CountingBatchExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )
            file_state = {
                "file_path": str(first_path),
                "file_paths": [str(first_path), str(second_path)],
                "preview_file_path": str(first_path),
                "sheet_name": "Main",
                "header_row": 1,
                "summary_column": "Summary",
                "preview_data": preview,
            }

            mapping_context = service.prepare_mapping_context(settings, file_state)
            self.assertEqual(
                CountingBatchExcelReader.read_excel_calls,
                [str(first_path), str(second_path)],
            )

            service.validate_mapping(
                mapping_context,
                mapping_context.selected_mapping,
            )
            result = service.run_batch_upload(
                settings,
                file_state,
                mapping_context,
                dry_run=True,
                continue_on_error=True,
                output_dir=str(Path(tmp_dir) / "output"),
            )

            self.assertEqual(
                CountingBatchExcelReader.read_excel_calls,
                [str(first_path), str(second_path)],
            )
            self.assertEqual(len(result["success_df"]), 4)

    def test_build_user_issue_df_keeps_only_user_visible_issues(self) -> None:
        service = GuiUploadPipelineService(client_factory=FakeClient)
        comparison_df = pd.DataFrame([
            {"df_column": "_row_id", "selected_schema_field": "", "status": MappingStatus.UNMAPPED.value},
            {"df_column": "담당자", "selected_schema_field": "", "status": MappingStatus.UNMAPPED.value},
        ])
        option_check_df = pd.DataFrame([
            {"df_column": "담당자", "schema_field": "담당자", "_row_id": 1, "raw_value": "홍길동", "status": "USER_NOT_FOUND"},
            {"df_column": "상태", "schema_field": "Status", "status": "PRECONSTRUCTION_REQUIRED"},
        ])
        payload_df = pd.DataFrame([
            {
                "_row_id": 1,
                "upload_name": "REQ-001",
                "payload_status": PayloadStatus.FAILED.value,
                "payload_error": "payload error",
            },
        ])
        row_context_df = pd.DataFrame([
            {
                "_row_id": 1,
                "_excel_row": 2,
                "upload_name": "REQ-001",
                "담당자": "홍길동",
            }
        ])

        visible_comparison_df = service._gui_visible_comparison_df(comparison_df)
        issue_df = service._build_user_issue_df(
            visible_comparison_df,
            option_check_df,
            payload_df,
            row_context_df=row_context_df,
        )

        self.assertFalse((issue_df["column"] == "_row_id").any())
        self.assertTrue((issue_df["column"] == "담당자").any())
        self.assertTrue(issue_df["message"].str.contains("사용자를 찾지 못했습니다").any())
        self.assertFalse(issue_df["message"].str.contains("내부 변환").any())
        self.assertTrue(issue_df["message"].str.contains("payload error").any())
        self.assertTrue((issue_df["row_label"] == "Excel 2행").any())
        self.assertTrue((issue_df["item_name"] == "REQ-001").any())
        self.assertTrue((issue_df["raw_value"] == "홍길동").any())
        self.assertTrue(issue_df["action"].str.contains("다시 검증").any())

    def test_build_user_issue_df_formats_structured_payload_errors(self) -> None:
        service = GuiUploadPipelineService(client_factory=FakeClient)
        issue_df = service._build_user_issue_df(
            pd.DataFrame(),
            pd.DataFrame(),
            pd.DataFrame([
                {
                    "_row_id": 1,
                    "upload_name": "REQ-001",
                    "payload_status": PayloadStatus.FAILED.value,
                    "payload_error": "[LOOKUP_REQUIRED] field='담당자' df_column='담당자' _row_id=1 lookup_target='user'",
                }
            ]),
            row_context_df=pd.DataFrame([
                {
                    "_row_id": 1,
                    "_excel_row": 2,
                    "upload_name": "REQ-001",
                    "담당자": "홍길동",
                }
            ]),
        )

        self.assertEqual(issue_df.iloc[0]["column"], "담당자")
        self.assertEqual(issue_df.iloc[0]["field"], "담당자")
        self.assertIn("필요한 값을 찾지 못해 업로드용 데이터를 만들 수 없습니다.", issue_df.iloc[0]["message"])
        self.assertEqual(issue_df.iloc[0]["row_label"], "Excel 2행")
        self.assertEqual(issue_df.iloc[0]["raw_value"], "홍길동")
        self.assertIn("다시 검증하세요", issue_df.iloc[0]["action"])

    def test_validate_mapping_does_not_block_when_payload_is_ready(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary", "담당자", "비고"])
            sheet.append(["REQ-001", "홍길동", "메모"])
            workbook.save(path)
            workbook.close()

            service = GuiUploadPipelineService(
                client_factory=FakeClient,
                excel_service=GuiExcelService(reader_cls=FakeExcelReader),
                reader_cls=FakeExcelReader,
            )
            settings = GuiSettings(
                base_url="https://example.com/cb",
                username="user",
                password="secret",
                default_project_id="10",
                default_tracker_id="1000",
                excel_header_row=1,
                summary_column="Summary",
                excel_sheet_name="Main",
            )
            mapping_context = service.prepare_mapping_context(
                settings,
                {
                    "file_path": str(path),
                    "sheet_name": "Main",
                    "header_row": 1,
                    "summary_column": "Summary",
                },
            )

            validation_context = service.validate_mapping(
                mapping_context,
                {"Summary": "Summary"},
                {"Status": "Open"},
            )

            self.assertFalse(validation_context.has_blocking_issues)
            self.assertTrue(validation_context.issue_df.empty)
            self.assertEqual(validation_context.summary_stats["total_rows"], 1)
            self.assertEqual(validation_context.summary_stats["ready_rows"], 1)
            self.assertEqual(validation_context.summary_stats["error_rows"], 0)
            self.assertEqual(
                list(mapping_context.wizard.state.payload_df["payload_status"]),
                [PayloadStatus.READY.value],
            )
            self.assertEqual(
                mapping_context.wizard.state.payload_df.iloc[0]["payload_json"]["status"]["name"],
                "Open",
            )
