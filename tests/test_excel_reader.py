from __future__ import annotations

from copy import copy
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from src.excel_reader import ExcelReader


class ExcelReaderTest(unittest.TestCase):
    def test_openpyxl_path_reads_sheet_names_headers_rows_and_indent(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["메타", "메타"])
            sheet.append(["Summary", "설명"])
            sheet.append(["Parent", "상위"])
            sheet.append([None, "연속 행"])
            sheet.append(["Child", "하위"])
            parent_alignment = copy(sheet["A3"].alignment)
            parent_alignment.indent = 1
            sheet["A3"].alignment = parent_alignment
            child_alignment = copy(sheet["A5"].alignment)
            child_alignment.indent = 2
            sheet["A5"].alignment = child_alignment
            workbook.create_sheet("Other")
            workbook.save(path)
            workbook.close()

            reader = ExcelReader(header_row=2, summary_col="Summary")

            self.assertEqual(reader.list_sheet_names(str(path)), ["Main", "Other"])
            self.assertEqual(reader.read_headers(str(path), "0"), ["Summary", "설명"])

            raw_df = reader.read_excel(str(path), sheet_name=0)

            self.assertEqual(raw_df["Summary"].tolist(), ["Parent", None, "Child"])
            self.assertEqual(raw_df["_excel_row"].tolist(), [3, 4, 5])
            self.assertEqual(raw_df["_summary_indent"].tolist(), [1, 0, 2])

    def test_count_upload_rows_supports_sheet_index_string(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            path = Path(tmp_dir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Main"
            sheet.append(["Summary", "설명"])
            sheet.append(["REQ-001", "첫 번째"])
            sheet.append([None, "연속 행"])
            sheet.append(["REQ-002", "두 번째"])
            workbook.save(path)
            workbook.close()

            reader = ExcelReader(header_row=1, summary_col="Summary")

            self.assertEqual(reader.count_upload_rows(str(path), "0"), 2)


if __name__ == "__main__":
    unittest.main()
