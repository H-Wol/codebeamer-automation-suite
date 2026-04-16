from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Callable
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from src.codebeamer_client import CodebeamerClient
from src.excel_reader import ExcelReader
from src.hierarchy_processor import HierarchyProcessor
from src.mapping_service import MappingService
from src.models import OptionCheckStatus
from src.wizard import CodebeamerUploadWizard


@dataclass
class PreviewData:
    sheet_names: list[str]
    headers: list[str]
    rows: list[list[str]]
    suggested_summary: str


class GuiCodebeamerService:
    """GUI 에서 사용하는 최소 Codebeamer 조회 기능을 제공한다."""

    def __init__(self, client_factory=CodebeamerClient, logger=None) -> None:
        self.client_factory = client_factory
        self.logger = logger

    def _build_client(self, settings) -> CodebeamerClient:
        return self.client_factory(
            settings.base_url,
            settings.username,
            settings.password,
            self.logger,
            rate_limit_retry_delay_seconds=settings.rate_limit_retry_delay_seconds,
            rate_limit_max_retries=settings.rate_limit_max_retries,
        )

    @staticmethod
    def _normalize_items(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        normalized: list[dict[str, Any]] = []
        for item in items:
            if not isinstance(item, dict):
                continue
            normalized.append({
                "id": item.get("id"),
                "name": item.get("name") or item.get("key") or str(item.get("id", "")),
            })
        return normalized

    def test_connection_and_load_projects(self, settings) -> list[dict[str, Any]]:
        projects = self._build_client(settings).get_projects()
        return self._normalize_items(projects)

    def load_trackers(self, settings, project_id: int) -> list[dict[str, Any]]:
        trackers = self._build_client(settings).get_trackers(project_id)
        return self._normalize_items(trackers)


class GuiExcelService:
    """GUI 파일 선택 화면에서 사용하는 Excel 메타데이터와 미리보기를 제공한다."""

    def __init__(self, logger=None) -> None:
        self.logger = logger

    @staticmethod
    def _normalize_headers(values: list[Any]) -> list[str]:
        headers: list[str] = []
        for index, value in enumerate(values):
            if value is None:
                headers.append(f"Unnamed_{index}")
            else:
                headers.append(str(value).strip())
        return headers

    @staticmethod
    def _suggest_summary(headers: list[str]) -> str:
        for header in headers:
            if header.lower() == "summary":
                return header
        for header in headers:
            if header == "요약":
                return header
        return headers[0] if headers else "Summary"

    def load_preview(
        self,
        file_path: str,
        *,
        sheet_name: str | None = None,
        header_row: int = 1,
        max_preview_rows: int = 10,
    ) -> PreviewData:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel 파일을 찾을 수 없습니다: {file_path}")
        if header_row < 1:
            raise ValueError("header_row 는 1 이상이어야 합니다.")

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet_names = workbook.sheetnames
            if not sheet_names:
                raise ValueError("시트가 없는 Excel 파일입니다.")

            target_sheet_name = sheet_name or sheet_names[0]
            if target_sheet_name not in sheet_names:
                target_sheet_name = sheet_names[0]

            sheet = workbook[target_sheet_name]
            rows_iter = sheet.iter_rows(values_only=True)
            rows = list(rows_iter)
            if len(rows) < header_row:
                headers: list[str] = []
                preview_rows: list[list[str]] = []
            else:
                headers = self._normalize_headers(list(rows[header_row - 1]))
                preview_rows = []
                for raw_row in rows[header_row:header_row + max_preview_rows]:
                    normalized = list(raw_row)
                    if len(normalized) < len(headers):
                        normalized += [None] * (len(headers) - len(normalized))
                    preview_rows.append(
                        ["" if value is None else str(value) for value in normalized[:len(headers)]]
                    )
            return PreviewData(
                sheet_names=sheet_names,
                headers=headers,
                rows=preview_rows,
                suggested_summary=self._suggest_summary(headers),
            )
        finally:
            workbook.close()

    def read_raw_dataframe(
        self,
        file_path: str,
        *,
        sheet_name: str | None = None,
        header_row: int = 1,
        summary_col: str = "Summary",
    ) -> pd.DataFrame:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel 파일을 찾을 수 없습니다: {file_path}")
        workbook = load_workbook(path, read_only=False, data_only=True)
        try:
            target_sheet_name = sheet_name or workbook.sheetnames[0]
            if target_sheet_name not in workbook.sheetnames:
                target_sheet_name = workbook.sheetnames[0]

            sheet = workbook[target_sheet_name]
            rows = list(sheet.iter_rows())
            if len(rows) < header_row:
                return pd.DataFrame()

            headers = self._normalize_headers([cell.value for cell in rows[header_row - 1]])
            if summary_col not in headers:
                raise ValueError(f"'{summary_col}' 컬럼을 찾을 수 없습니다.")

            summary_index = headers.index(summary_col)
            records: list[dict[str, Any]] = []
            for row in rows[header_row:]:
                values = [cell.value for cell in row[:len(headers)]]
                if all(value is None or str(value).strip() == "" for value in values):
                    continue
                indent = 0
                if summary_index < len(row):
                    try:
                        indent = int(row[summary_index].alignment.indent or 0)
                    except Exception:
                        indent = 0
                record = {header: values[index] if index < len(values) else None for index, header in enumerate(headers)}
                record["_excel_row"] = row[0].row if row else None
                record["_summary_indent"] = indent
                records.append(record)
            return pd.DataFrame(records)
        finally:
            workbook.close()


BLOCKING_OPTION_STATUSES = {
    OptionCheckStatus.DIRECT_PARSE_FAILED.value,
    OptionCheckStatus.DF_COLUMN_MISSING.value,
    OptionCheckStatus.FIELD_UNSUPPORTED.value,
    OptionCheckStatus.LOOKUP_REQUIRED.value,
    OptionCheckStatus.OPTION_MAP_MISSING.value,
    OptionCheckStatus.OPTION_NOT_FOUND.value,
    OptionCheckStatus.OPTION_SOURCE_UNAVAILABLE.value,
}
USER_LOOKUP_FAILURE_SUFFIXES = (
    "USER_LOOKUP_NOT_RUN",
    "USER_LOOKUP_FAILED",
    "USER_LOOKUP_AMBIGUOUS",
    "USER_NOT_FOUND",
    "MEMBER_LOOKUP_FAILED",
    "MEMBER_LOOKUP_AMBIGUOUS",
    "MEMBER_NOT_FOUND",
)
GUI_EXCLUDED_MAPPING_COLUMNS = {
    "id",
    "parent",
    "parent_row_id",
    "upload_name",
    "depth",
    "_row_id",
    "_summary_indent",
    "_start_excel_row",
    "_end_excel_row",
    "_excel_row",
}


@dataclass
class MappingContext:
    wizard: CodebeamerUploadWizard
    schema_df: pd.DataFrame
    upload_columns: list[str]
    selected_mapping: dict[str, str]
    list_cols: list[str]


@dataclass
class ValidationContext:
    comparison_df: pd.DataFrame
    option_check_df: pd.DataFrame
    converted_upload_df: pd.DataFrame
    has_blocking_issues: bool


class GuiUploadPipelineService:
    """GUI 단계가 재사용할 업로드 파이프라인 래퍼다."""

    def __init__(self, logger=None, *, client_factory=CodebeamerClient) -> None:
        self.logger = logger
        self.mapper = MappingService(logger=logger)
        self.client_factory = client_factory

    def create_wizard(self, settings) -> CodebeamerUploadWizard:
        client = self.client_factory(
            settings.base_url,
            settings.username,
            settings.password,
            self.logger,
            rate_limit_retry_delay_seconds=settings.rate_limit_retry_delay_seconds,
            rate_limit_max_retries=settings.rate_limit_max_retries,
        )
        reader = ExcelReader(
            header_row=settings.excel_header_row,
            summary_col=settings.summary_column,
            logger=self.logger,
        )
        processor = HierarchyProcessor(
            header_row=settings.excel_header_row,
            summary_col=settings.summary_column,
            logger=self.logger,
        )
        return CodebeamerUploadWizard(
            client=client,
            processor=processor,
            mapper=self.mapper,
            reader=reader,
            logger=self.logger,
        )

    @staticmethod
    def _auto_match_columns(upload_columns: list[str], schema_field_names: set[str]) -> dict[str, str]:
        selected_mapping: dict[str, str] = {}
        for column in upload_columns:
            if column in schema_field_names:
                selected_mapping[column] = column
                continue
            for schema_field in schema_field_names:
                if column.lower() == schema_field.lower():
                    selected_mapping[column] = schema_field
                    break
        return selected_mapping

    @staticmethod
    def _gui_upload_columns(upload_df: pd.DataFrame) -> list[str]:
        columns: list[str] = []
        for column in upload_df.columns:
            if column in GUI_EXCLUDED_MAPPING_COLUMNS:
                continue
            if str(column).startswith("_"):
                continue
            columns.append(str(column))
        return columns

    def prepare_mapping_context(self, settings, file_state: dict[str, Any]) -> MappingContext:
        wizard = self.create_wizard(settings)
        wizard.select_project(int(settings.default_project_id))
        wizard.select_tracker(int(settings.default_tracker_id))

        schema = wizard.client.get_tracker_schema(wizard.state.tracker_id)
        schema_df = self.mapper.flatten_schema_fields(schema)

        preview = GuiExcelService(logger=self.logger).load_preview(
            file_state["file_path"],
            sheet_name=file_state["sheet_name"],
            header_row=int(file_state["header_row"]),
        )
        headers = preview.headers
        raw_mapping = self._auto_match_columns(
            headers,
            set(schema_df["field_name"].dropna().astype(str).str.strip()),
        )
        list_cols = self.mapper.get_list_columns_for_mapping(raw_mapping, schema_df)

        raw_df = GuiExcelService(logger=self.logger).read_raw_dataframe(
            file_path=file_state["file_path"],
            sheet_name=file_state["sheet_name"],
            header_row=int(file_state["header_row"]),
            summary_col=str(file_state["summary_column"]),
        )
        wizard.load_raw_dataframe(raw_df, list_cols=list_cols)
        wizard.state.schema = schema
        wizard.state.schema_df = schema_df
        wizard._detect_table_field_columns()

        upload_columns = self._gui_upload_columns(wizard.state.upload_df)
        selected_mapping = {
            column: schema_field
            for column, schema_field in raw_mapping.items()
            if column in upload_columns
        }

        return MappingContext(
            wizard=wizard,
            schema_df=schema_df,
            upload_columns=upload_columns,
            selected_mapping=selected_mapping,
            list_cols=list_cols,
        )

    def validate_mapping(
        self,
        mapping_context: MappingContext,
        selected_mapping: dict[str, str],
    ) -> ValidationContext:
        wizard = mapping_context.wizard
        list_cols = self.mapper.get_list_columns_for_mapping(selected_mapping, mapping_context.schema_df)
        wizard.load_raw_dataframe(wizard.state.raw_df.copy(), list_cols=list_cols)
        comparison_df = wizard.load_schema_and_compare(selected_mapping)
        _, option_check_df = wizard.process_option_mapping(selected_mapping)
        option_check_df = option_check_df if option_check_df is not None else pd.DataFrame()

        has_blocking = False
        if not option_check_df.empty:
            status_series = option_check_df["status"].fillna("").astype(str)
            has_blocking = bool(
                status_series.isin(BLOCKING_OPTION_STATUSES).any()
                or status_series.str.endswith(USER_LOOKUP_FAILURE_SUFFIXES).any()
            )

        payload_df = wizard.build_payloads(force=True)
        if not payload_df.empty and (payload_df["payload_status"] != "ready").any():
            has_blocking = True

        return ValidationContext(
            comparison_df=comparison_df,
            option_check_df=option_check_df,
            converted_upload_df=wizard.state.converted_upload_df,
            has_blocking_issues=has_blocking,
        )
