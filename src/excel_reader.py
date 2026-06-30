from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


class ExcelReader:
    """Excel 파일에서 raw DataFrame만 읽어오는 입력 전용 reader다."""

    def __init__(self, header_row: int = 1, summary_col: str = "Summary", logger=None):
        self.header_row = header_row
        self.summary_col = summary_col
        self.logger = logger

    @staticmethod
    def is_blank(value: Any) -> bool:
        if value is None:
            return True
        if isinstance(value, float) and pd.isna(value):
            return True
        if isinstance(value, str) and value.strip() == "":
            return True
        return False

    @staticmethod
    def _normalize_headers(headers: list[Any]) -> list[str]:
        return [
            str(header).strip() if header is not None else f"Unnamed_{index}"
            for index, header in enumerate(headers)
        ]

    @staticmethod
    def _supports_openpyxl(file_path: str) -> bool:
        return Path(file_path).suffix.lower() in {".xlsx", ".xlsm", ".xltx", ".xltm"}

    @staticmethod
    def _resolve_openpyxl_sheet(workbook, sheet_name: str | int):
        if isinstance(sheet_name, str):
            normalized = sheet_name.strip()
            if normalized in workbook.sheetnames:
                return workbook[normalized]
            if normalized.isdigit():
                sheet_index = int(normalized)
                if 0 <= sheet_index < len(workbook.worksheets):
                    return workbook.worksheets[sheet_index]
        elif isinstance(sheet_name, int) and 0 <= sheet_name < len(workbook.worksheets):
            return workbook.worksheets[sheet_name]

        raise ValueError(f"시트를 찾을 수 없습니다: {sheet_name}")

    @staticmethod
    def _normalize_row(values: list[Any], width: int) -> list[Any]:
        normalized = list(values)
        if len(normalized) < width:
            normalized += [None] * (width - len(normalized))
        elif len(normalized) > width:
            normalized = normalized[:width]
        return normalized

    def _openpyxl_sheet_names(self, file_path: str) -> list[str]:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            return list(workbook.sheetnames)
        finally:
            workbook.close()

    def _openpyxl_headers(self, file_path: str, sheet_name: str | int) -> list[str]:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            worksheet = self._resolve_openpyxl_sheet(workbook, sheet_name)
            row = next(
                worksheet.iter_rows(
                    min_row=self.header_row,
                    max_row=self.header_row,
                ),
                (),
            )
            return self._normalize_headers([cell.value for cell in row])
        finally:
            workbook.close()

    @staticmethod
    def _create_xlwings_app(visible: bool = False):
        try:
            import xlwings as xw
        except ImportError as exc:  # pragma: no cover
            raise RuntimeError("Excel 구형 포맷 처리에는 xlwings 패키지가 필요합니다.") from exc

        app = xw.App(visible=visible, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        return app

    @staticmethod
    def _resolve_xlwings_sheet(workbook, sheet_name: str | int):
        if isinstance(sheet_name, str):
            normalized = sheet_name.strip()
            if normalized.isdigit():
                sheet_index = int(normalized)
                return workbook.sheets[sheet_index]
            return workbook.sheets[normalized]
        return workbook.sheets[sheet_name]

    def _xlwings_sheet_names(self, file_path: str) -> list[str]:
        app = self._create_xlwings_app()
        workbook = None

        try:
            workbook = app.books.open(file_path)
            return [sheet.name for sheet in workbook.sheets]
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass
            app.quit()

    def _xlwings_headers(self, file_path: str, sheet_name: str | int) -> list[str]:
        app = self._create_xlwings_app()
        workbook = None

        try:
            workbook = app.books.open(file_path)
            sheet = self._resolve_xlwings_sheet(workbook, sheet_name)
            values = sheet.used_range.value
            if not values or len(values) < self.header_row:
                return []
            headers = values[self.header_row - 1]
            return self._normalize_headers(headers)
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass
            app.quit()

    def list_sheet_names(self, file_path: str) -> list[str]:
        if self._supports_openpyxl(file_path):
            return self._openpyxl_sheet_names(file_path)
        return self._xlwings_sheet_names(file_path)

    def read_headers(self, file_path: str, sheet_name: str | int) -> list[str]:
        if self._supports_openpyxl(file_path):
            return self._openpyxl_headers(file_path, sheet_name)
        return self._xlwings_headers(file_path, sheet_name)

    def count_upload_rows(self, file_path: str, sheet_name: str | int) -> int:
        if self._supports_openpyxl(file_path):
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            try:
                worksheet = self._resolve_openpyxl_sheet(workbook, sheet_name)
                header_row = next(
                    worksheet.iter_rows(
                        min_row=self.header_row,
                        max_row=self.header_row,
                    ),
                    (),
                )
                headers = self._normalize_headers([cell.value for cell in header_row])
                if self.summary_col not in headers:
                    raise ValueError(f"'{self.summary_col}' 컬럼을 찾을 수 없습니다.")
                summary_index = headers.index(self.summary_col)
                upload_row_count = 0
                for row in worksheet.iter_rows(min_row=self.header_row + 1):
                    normalized_row = self._normalize_row([cell.value for cell in row], len(headers))

                    if all(self.is_blank(value) for value in normalized_row):
                        continue
                    if not self.is_blank(normalized_row[summary_index]):
                        upload_row_count += 1
                return upload_row_count
            finally:
                workbook.close()

        headers = self.read_headers(file_path, sheet_name)
        if self.summary_col not in headers:
            raise ValueError(f"'{self.summary_col}' 컬럼을 찾을 수 없습니다.")
        raw_df = self.read_excel(file_path=file_path, sheet_name=sheet_name)
        if raw_df.empty:
            return 0
        return int((~raw_df[self.summary_col].apply(self.is_blank)).sum())

    def _resolve_indent_level(self, cell: Any, summary_value: Any) -> int:
        try:
            alignment = getattr(cell, "alignment", None)
            indent = getattr(alignment, "indent", None) if alignment is not None else None
            if indent not in (None, ""):
                return int(indent)
        except Exception:
            pass
        try:
            return int(cell.api.IndentLevel)
        except Exception:
            if isinstance(summary_value, str):
                leading_spaces = len(summary_value) - len(summary_value.lstrip(" "))
                return leading_spaces // 4
        return 0

    def read_excel(self, file_path: str, sheet_name: str | int = 0, visible: bool = False) -> pd.DataFrame:
        """Excel 시트를 읽고 `_excel_row`, `_summary_indent` 메타정보를 포함한 raw DataFrame을 만든다."""
        if self._supports_openpyxl(file_path):
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            try:
                worksheet = self._resolve_openpyxl_sheet(workbook, sheet_name)
                header_cells = next(
                    worksheet.iter_rows(
                        min_row=self.header_row,
                        max_row=self.header_row,
                    ),
                    (),
                )
                headers = self._normalize_headers([cell.value for cell in header_cells])
                if self.summary_col not in headers:
                    raise ValueError(f"'{self.summary_col}' 컬럼을 찾을 수 없습니다.")

                summary_col_index = headers.index(self.summary_col)
                records = []
                for excel_row, row in enumerate(
                    worksheet.iter_rows(min_row=self.header_row + 1),
                    start=self.header_row + 1,
                ):
                    normalized_row = self._normalize_row([cell.value for cell in row], len(headers))
                    if all(self.is_blank(value) for value in normalized_row):
                        continue

                    summary_cell = row[summary_col_index] if summary_col_index < len(row) else None
                    summary_value = normalized_row[summary_col_index]
                    indent_level = self._resolve_indent_level(summary_cell, summary_value)

                    record = dict(zip(headers, normalized_row))
                    record["_excel_row"] = excel_row
                    record["_summary_indent"] = indent_level
                    records.append(record)

                return pd.DataFrame(records)
            finally:
                workbook.close()

        app = self._create_xlwings_app(visible=visible)
        workbook = None

        try:
            workbook = app.books.open(file_path)
            sheet = self._resolve_xlwings_sheet(workbook, sheet_name)
            used_range = sheet.used_range
            values = used_range.value

            headers = self._normalize_headers(values[self.header_row - 1])
            if self.summary_col not in headers:
                raise ValueError(f"'{self.summary_col}' 컬럼을 찾을 수 없습니다.")

            summary_col_idx_1based = headers.index(self.summary_col) + 1
            data_rows = values[self.header_row:]
            used_start_row = used_range.row
            used_start_col = used_range.column

            records = []
            for relative_index, row in enumerate(data_rows, start=self.header_row + 1):
                normalized_row = self._normalize_row(
                    list(row) if isinstance(row, list) else [row],
                    len(headers),
                )

                if all(self.is_blank(value) for value in normalized_row):
                    continue

                excel_row = used_start_row + (relative_index - 1)
                excel_col = used_start_col + (summary_col_idx_1based - 1)
                summary_cell = sheet.range((excel_row, excel_col))
                summary_value = normalized_row[summary_col_idx_1based - 1]
                indent_level = self._resolve_indent_level(summary_cell, summary_value)

                record = dict(zip(headers, normalized_row))
                record["_excel_row"] = excel_row
                record["_summary_indent"] = indent_level
                records.append(record)

            return pd.DataFrame(records)
        finally:
            if workbook is not None:
                try:
                    workbook.close()
                except Exception:
                    pass
            app.quit()
